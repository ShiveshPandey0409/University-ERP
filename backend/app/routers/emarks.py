"""Marks entry (Emarks). Reads: any authenticated; mark correction: role 5 or 12."""
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.core.deps import Principal, get_current_user, require_roles
from app.db.session import get_db
from app.schemas.exam import EntryStatusItem, MarkRow, MarkUpdateRequest, PaperItem
from app.services import emarks_service

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

router = APIRouter(prefix="/emarks", tags=["emarks"])


@router.get("/entry-status", response_model=list[EntryStatusItem])
def entry_status(
    course_id: str | None = Query(default=None),
    semester: str | None = Query(default=None),
    db: Session = Depends(get_db),
    _: Principal = Depends(get_current_user),
):
    return emarks_service.entry_status(db, course_id=course_id, semester=semester)


@router.get("/papers", response_model=list[PaperItem])
def papers(course_id: str, semester: str, db: Session = Depends(get_db),
           _: Principal = Depends(get_current_user)):
    return emarks_service.paper_list(db, course_id=course_id, semester=semester)


@router.get("/marks", response_model=list[MarkRow])
def marks(course_id: str, semester: str, paper_code: str, paper_type: str,
          db: Session = Depends(get_db), _: Principal = Depends(get_current_user)):
    return emarks_service.marks_by_paper(
        db, course_id=course_id, semester=semester, paper_code=paper_code, paper_type=paper_type
    )


@router.get("/search/{rollno}")
def search(rollno: str, db: Session = Depends(get_db), _: Principal = Depends(get_current_user)) -> list[dict]:
    return emarks_service.search_by_roll(db, rollno)


@router.post("/marks/{mark_id}")
def update_mark(mark_id: int, body: MarkUpdateRequest, db: Session = Depends(get_db),
                user: Principal = Depends(require_roles(5, 12))):
    updated = emarks_service.update_mark(db, mark_id=mark_id, marks=body.marks, by=user.uname)
    if updated == 0:
        raise HTTPException(404, "Mark row not found")
    return {"updated": True, "id": mark_id, "marks": body.marks}


@router.post("/upload")
async def upload_marks(
    course_id: str, semester: str, paper_code: str, paper_type: str,
    dry_run: bool = Query(default=True, description="true = validate only, no DB writes"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: Principal = Depends(require_roles(5, 12)),
):
    """Bulk marks upload from an .xlsx with 'rollno' and 'mark' columns."""
    wb = load_workbook(BytesIO(await file.read()), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = [str(c).strip().lower() if c is not None else "" for c in (next(it, []) or [])]

    def col(*names):
        for i, h in enumerate(header):
            if h in names:
                return i
        return None

    ri = col("rollno", "roll no", "roll", "roll_no")
    mi = col("mark", "marks", "mark1", "obtained")
    if ri is None or mi is None:
        raise HTTPException(400, f"Sheet needs 'rollno' and 'mark' columns; found: {header}")

    rows = []
    for r in it:
        if not r or (ri < len(r) and r[ri] is None):
            continue
        roll = r[ri] if ri < len(r) else None
        mark = r[mi] if mi < len(r) else None
        if roll is None:
            continue
        rows.append({"rollno": str(roll).strip(), "mark": "" if mark is None else str(mark).strip()})

    return emarks_service.apply_bulk_marks(
        db, course_id=course_id, semester=semester, paper_code=paper_code,
        paper_type=paper_type, rows=rows, dry_run=dry_run, by=user.uname,
    )


@router.get("/marks.xlsx")
def marks_xlsx(course_id: str, semester: str, paper_code: str, paper_type: str,
               db: Session = Depends(get_db), _: Principal = Depends(get_current_user)):
    """Mark-list Excel export (replaces the legacy ClosedXML MrkListXls)."""
    rows = emarks_service.marks_by_paper(
        db, course_id=course_id, semester=semester, paper_code=paper_code, paper_type=paper_type
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Marks"
    ws.append(["Roll No", "Enrollment", "Name", "Max Marks", "Marks"])
    for m in rows:
        ws.append([m["rollno"], m["enroll_no"], m["name"], m["mm"], m["mark1"]])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"marks_{course_id}_{semester}_{paper_code}.xlsx"
    return StreamingResponse(
        buf, media_type=XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
